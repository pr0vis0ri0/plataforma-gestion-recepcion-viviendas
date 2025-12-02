import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from core.utils.region_metrics import get_region_metrics
from core.utils.cumplimiento_constructora import get_cumplimiento_plazos_por_constructora
from proyectos.models import Proyecto, Vivienda
from incidencias.models import Observacion
from core.models import Region
from django.db.models import Count, Q, F, ExpressionWrapper, DurationField, Avg
from django.db.models.functions import TruncMonth
from django.contrib.auth import get_user_model
from datetime import datetime

@login_required
def dashboard_excel_report(request):
    region_id = request.GET.get('region')
    estado_id = request.GET.get('estado')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    metrics_region = get_region_metrics(region_id=region_id, estado=estado_id, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    cumplimiento_constructoras = get_cumplimiento_plazos_por_constructora(region_id=region_id, estado=estado_id, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)

    proyectos_qs = Proyecto.objects.filter(activo=True)
    viviendas_qs = Vivienda.objects.filter(activa=True)
    obs_qs = Observacion.objects.filter(activo=True)
    if region_id:
        proyectos_qs = proyectos_qs.filter(region_id=region_id)
        viviendas_qs = viviendas_qs.filter(proyecto__region_id=region_id)
        obs_qs = obs_qs.filter(vivienda__proyecto__region_id=region_id)
    if fecha_inicio:
        proyectos_qs = proyectos_qs.filter(fecha_creacion__date__gte=fecha_inicio)
        viviendas_qs = viviendas_qs.filter(proyecto__fecha_creacion__date__gte=fecha_inicio)
        obs_qs = obs_qs.filter(fecha_creacion__date__gte=fecha_inicio)
    if fecha_fin:
        proyectos_qs = proyectos_qs.filter(fecha_creacion__date__lte=fecha_fin)
        viviendas_qs = viviendas_qs.filter(proyecto__fecha_creacion__date__lte=fecha_fin)
        obs_qs = obs_qs.filter(fecha_creacion__date__lte=fecha_fin)

    # 1. KPIs principales
    wb = openpyxl.Workbook()
    ws_kpi = wb.active
    ws_kpi.title = "KPIs"
    kpi_headers = ["Total Viviendas", "Viviendas Entregadas", "% Viviendas Entregadas", "Casos Postventa Abiertos", "Tiempo Promedio Resolución (días)", "Familias Acompañadas", "% Familias Acompañadas", "Tasa Cumplimiento"]
    ws_kpi.append(kpi_headers)
    viviendas_total = viviendas_qs.count()
    viviendas_entregadas = viviendas_qs.filter(estado='entregada').count()
    porc_viviendas_entregadas = round((viviendas_entregadas / viviendas_total) * 100, 1) if viviendas_total else 0
    casos_postventa_abiertos = obs_qs.filter(estado__nombre='Abierta').count()
    obs_cerradas = obs_qs.filter(estado__nombre='Cerrada', fecha_cierre__isnull=False, fecha_creacion__isnull=False)
    if obs_cerradas.exists():
        expr = ExpressionWrapper(F('fecha_cierre') - F('fecha_creacion'), output_field=DurationField())
        promedio = obs_cerradas.annotate(duracion=expr).aggregate(prom=Avg('duracion'))['prom']
        tiempo_promedio_resolucion = int(promedio.total_seconds() // 86400) if promedio else 0
    else:
        tiempo_promedio_resolucion = 0
    familias_acompañadas = viviendas_qs.exclude(beneficiario__isnull=True).count()
    porc_familias_acompañadas = round((familias_acompañadas / viviendas_total) * 100, 1) if viviendas_total else 0
    tasa_cumplimiento = cumplimiento_constructoras['global']
    ws_kpi.append([
        viviendas_total, viviendas_entregadas, porc_viviendas_entregadas, casos_postventa_abiertos,
        tiempo_promedio_resolucion, familias_acompañadas, porc_familias_acompañadas, tasa_cumplimiento
    ])
    for col in range(1, len(kpi_headers)+1):
        ws_kpi.column_dimensions[get_column_letter(col)].width = 22
    for cell in ws_kpi[1]:
        cell.font = Font(bold=True)

    # 2. Estados de Vivienda
    ws_estado = wb.create_sheet("Estados Vivienda")
    ws_estado.append(["Estado", "Cantidad", "Porcentaje"])
    estados_vivienda = viviendas_qs.values('estado').annotate(cantidad=Count('id')).order_by('-cantidad')
    for ev in estados_vivienda:
        nombre = ev['estado'].capitalize() if ev['estado'] else 'Sin estado'
        cantidad = ev['cantidad']
        porcentaje = round((cantidad / viviendas_total) * 100, 1) if viviendas_total else 0
        ws_estado.append([nombre, cantidad, porcentaje])
    for cell in ws_estado[1]:
        cell.font = Font(bold=True)

    # 3. Observaciones por tipo
    ws_obs = wb.create_sheet("Observaciones")
    ws_obs.append(["Tipo de Observación", "Totales", "Cerrados", "Pendientes", "Tiempo Promedio (días)"])
    tipos_obs = obs_qs.values('tipo__nombre').annotate(
        totales=Count('id'),
        cerrados=Count('id', filter=Q(estado__nombre='Cerrada')),
        pendientes=Count('id', filter=Q(estado__nombre='Abierta')),
        tiempo_promedio=Avg(ExpressionWrapper(F('fecha_cierre') - F('fecha_creacion'), output_field=DurationField()), filter=Q(estado__nombre='Cerrada', fecha_cierre__isnull=False, fecha_creacion__isnull=False))
    ).order_by('-totales')
    for t in tipos_obs:
        if t['tiempo_promedio']:
            dias = int(t['tiempo_promedio'].total_seconds() // 86400)
        else:
            dias = '-'
        ws_obs.append([
            t['tipo__nombre'] or 'Sin tipo', t['totales'], t['cerrados'], t['pendientes'], dias
        ])
    for cell in ws_obs[1]:
        cell.font = Font(bold=True)

    # 4. Tendencias temporales
    ws_tend = wb.create_sheet("Tendencias")
    ws_tend.append(["Mes", "Abiertos", "Cerrados", "Variación"])
    obs_mes = obs_qs.annotate(mes=TruncMonth('fecha_creacion')).values('mes').annotate(
        abiertos=Count('id', filter=Q(estado__nombre='Abierta')),
        cerrados=Count('id', filter=Q(estado__nombre='Cerrada'))
    ).order_by('mes')
    prev_cerrados = 0
    for m in obs_mes:
        nombre = m['mes'].strftime('%B') if m['mes'] else 'Sin mes'
        abiertos = m['abiertos']
        cerrados = m['cerrados']
        variacion = cerrados - prev_cerrados
        ws_tend.append([nombre, abiertos, cerrados, variacion])
        prev_cerrados = cerrados
    for cell in ws_tend[1]:
        cell.font = Font(bold=True)

    # 5. Desempeño regional
    ws_reg = wb.create_sheet("Desempeño Regional")
    ws_reg.append(["Región", "Total", "Entregadas", "Casos", "Tiempo"])
    for r in metrics_region:
        if not region_id or (region_id and str(r.get('region_id', '')) == str(region_id)):
            ws_reg.append([
                r['region'], r['total_viviendas'], r['entregadas'], r['casos_postventa'], r['promedio_dias']
            ])
    for cell in ws_reg[1]:
        cell.font = Font(bold=True)

    # 6. Desempeño del equipo
    ws_eq = wb.create_sheet("Equipo")
    ws_eq.append(["Técnico", "Asignados", "Cerrados", "Tasa Cierre", "Tiempo Promedio"])
    User = get_user_model()
    tecnicos = User.objects.filter(is_active=True, rol__nombre__in=['TECNICO', 'COORDINADOR'])
    for t in tecnicos:
        asignados = obs_qs.filter(asignado_a=t).count()
        cerrados = obs_qs.filter(asignado_a=t, estado__nombre='Cerrada').count()
        tasa_cierre = round((cerrados / asignados) * 100, 1) if asignados else 0
        tiempo_promedio = '-'
        obs_cerradas = obs_qs.filter(asignado_a=t, estado__nombre='Cerrada', fecha_cierre__isnull=False, fecha_creacion__isnull=False)
        if obs_cerradas.exists():
            expr = ExpressionWrapper(F('fecha_cierre') - F('fecha_creacion'), output_field=DurationField())
            promedio = obs_cerradas.annotate(duracion=expr).aggregate(prom=Avg('duracion'))['prom']
            if promedio:
                tiempo_promedio = int(promedio.total_seconds() // 86400)
        ws_eq.append([
            t.get_full_name() if hasattr(t, 'get_full_name') else (t.first_name + ' ' + t.last_name).strip() or t.username,
            asignados, cerrados, tasa_cierre, tiempo_promedio
        ])
    for cell in ws_eq[1]:
        cell.font = Font(bold=True)


    # 7. Detalle de Viviendas
    ws_viv = wb.create_sheet("Detalle Viviendas")
    viv_headers = [
        "ID", "Código", "Proyecto", "Región", "Estado", "Beneficiario", "Fecha Creación", "Tipología"
    ]
    ws_viv.append(viv_headers)
    for v in viviendas_qs.select_related('proyecto', 'proyecto__region', 'beneficiario', 'tipologia'):
        ws_viv.append([
            v.id,
            v.codigo,
            v.proyecto.nombre if v.proyecto else '',
            v.proyecto.region.nombre if v.proyecto and v.proyecto.region else '',
            v.estado,
            f"{v.beneficiario.nombre} {v.beneficiario.apellido_paterno if hasattr(v.beneficiario, 'apellido_paterno') else ''}" if v.beneficiario else '',
            v.fecha_creacion.strftime('%Y-%m-%d') if v.fecha_creacion else '',
            v.tipologia.nombre if v.tipologia else '',
        ])
    for cell in ws_viv[1]:
        cell.font = Font(bold=True)

    # 8. Detalle de Observaciones
    ws_obsdet = wb.create_sheet("Detalle Observaciones")
    obs_headers = [
        "ID", "Vivienda", "Proyecto", "Región", "Tipo", "Estado", "Detalle", "Fecha Creación", "Fecha Cierre", "Asignado a"
    ]
    ws_obsdet.append(obs_headers)
    for o in obs_qs.select_related('vivienda', 'vivienda__proyecto', 'vivienda__proyecto__region', 'tipo', 'estado', 'asignado_a'):
        ws_obsdet.append([
            o.id,
            o.vivienda.codigo if o.vivienda else '',
            o.vivienda.proyecto.nombre if o.vivienda and o.vivienda.proyecto else '',
            o.vivienda.proyecto.region.nombre if o.vivienda and o.vivienda.proyecto and o.vivienda.proyecto.region else '',
            o.tipo.nombre if o.tipo else '',
            o.estado.nombre if o.estado else '',
            o.detalle,
            o.fecha_creacion.strftime('%Y-%m-%d') if o.fecha_creacion else '',
            o.fecha_cierre.strftime('%Y-%m-%d') if o.fecha_cierre else '',
            o.asignado_a.get_full_name() if o.asignado_a and hasattr(o.asignado_a, 'get_full_name') else (o.asignado_a.username if o.asignado_a else ''),
        ])
    for cell in ws_obsdet[1]:
        cell.font = Font(bold=True)

    # 9. Detalle de Proyectos
    ws_proj = wb.create_sheet("Detalle Proyectos")
    proj_headers = [
        "ID", "Nombre", "Región", "Comuna", "Fecha Creación"
    ]
    ws_proj.append(proj_headers)
    for p in proyectos_qs.select_related('region', 'comuna'):
        ws_proj.append([
            p.id,
            p.nombre,
            p.region.nombre if p.region else '',
            p.comuna.nombre if p.comuna else '',
            p.fecha_creacion.strftime('%Y-%m-%d') if p.fecha_creacion else '',
        ])
    for cell in ws_proj[1]:
        cell.font = Font(bold=True)

    # Guardar y responder
    fecha_hora = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f"dashboard_techoChile_{fecha_hora}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
