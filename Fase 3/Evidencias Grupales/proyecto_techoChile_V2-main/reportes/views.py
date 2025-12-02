from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import ReporteGenerado
from django.http import FileResponse, Http404
import os

# Vista para listar reportes generados
@login_required
def listar_reportes_generados(request):
    from django.db.models.functions import TruncDate
    # Obtener todas las fechas únicas (solo día)
    fechas_unicas = ReporteGenerado.objects.annotate(fecha_dia=TruncDate('fecha_generacion'))\
        .values_list('fecha_dia', flat=True).distinct().order_by('-fecha_dia')

    # Filtro por fecha (YYYY-MM-DD)
    fecha_filtro = request.GET.get('fecha')
    reportes = ReporteGenerado.objects.all()
    if fecha_filtro:
        # Buscar reportes de ese día
        reportes = reportes.filter(fecha_generacion__date=fecha_filtro)

    # Para cada reporte, calcular filtros válidos (sin 'estado' vacío)
    reportes_list = []
    from core.models import Region
    for r in reportes:
        filtros_validos = []
        if hasattr(r, 'filtros') and isinstance(r.filtros, dict):
            for k, v in r.filtros.items():
                # Ocultar 'estado' siempre
                if k == 'estado':
                    continue
                # Si es region y tiene valor, buscar el nombre
                if k == 'region' and v:
                    try:
                        region_obj = Region.objects.get(pk=int(v))
                        v = region_obj.nombre
                    except Exception:
                        pass
                filtros_validos.append((k, v))
        # Si los únicos filtros son fecha_inicio y/o fecha_fin y ambos vacíos, considerar como sin filtro
        solo_fechas = all(k in ['fecha_inicio', 'fecha_fin'] for k, v in filtros_validos)
        fechas_vacias = all(v == '' for k, v in filtros_validos if k in ['fecha_inicio', 'fecha_fin'])
        mostrar_sin_filtro = (filtros_validos == [] or (solo_fechas and fechas_vacias))
        reportes_list.append({
            'obj': r,
            'filtros_validos': filtros_validos,
            'mostrar_sin_filtro': mostrar_sin_filtro,
        })

    return render(request, 'reportes/listar_reportes_generados.html', {
        'reportes': reportes_list,
        'fechas_unicas': fechas_unicas,
        'fecha_filtro': fecha_filtro or '',
    })

# Vista para descargar un reporte generado
@login_required
def descargar_reporte_generado(request, reporte_id):
    reporte = get_object_or_404(ReporteGenerado, id=reporte_id)
    ruta = os.path.join(os.path.dirname(os.path.dirname(__file__)), reporte.ruta_archivo)
    if not os.path.exists(ruta):
        raise Http404('Archivo no encontrado')
    return FileResponse(open(ruta, 'rb'), as_attachment=True, filename=reporte.nombre_archivo)
from django.http import HttpResponse
import openpyxl
from django.contrib.auth.decorators import login_required

@login_required
def reporte_viviendas_sin_observaciones_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Viviendas sin Observaciones"
    headers = [
        "Proyecto", "Código Vivienda", "Tipología", "Estado", "Beneficiario", "RUT", "Email", "Constructora", "Región", "Comuna"
    ]
    ws.append(headers)
    from proyectos.models import Vivienda
    viviendas = Vivienda.objects.select_related('proyecto', 'beneficiario', 'tipologia', 'proyecto__constructora', 'proyecto__region', 'proyecto__comuna')\
        .filter(observaciones__isnull=True, activa=True)
    for vivienda in viviendas:
        beneficiario = vivienda.beneficiario
        proyecto = vivienda.proyecto
        ws.append([
            proyecto.nombre if proyecto else "",
            vivienda.codigo,
            vivienda.tipologia.nombre if vivienda.tipologia else "",
            vivienda.estado,
            beneficiario.nombre_completo if beneficiario else "",
            beneficiario.rut if beneficiario else "",
            beneficiario.email if beneficiario else "",
            proyecto.constructora.nombre if proyecto and proyecto.constructora else "",
            proyecto.region.nombre if proyecto and proyecto.region else "",
            proyecto.comuna.nombre if proyecto and proyecto.comuna else ""
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=viviendas_sin_observaciones.xlsx'
    wb.save(response)
    return response


# Reporte Total: Viviendas y Beneficiarios

@login_required
def reporte_total_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Total"
    headers = ["Proyecto", "Vivienda", "Tipología", "Estado Vivienda", "Beneficiario", "RUT", "Email"]
    ws.append(headers)
    from proyectos.models import Proyecto, Vivienda
    viviendas = Vivienda.objects.select_related('proyecto', 'beneficiario', 'tipologia').all()
    for vivienda in viviendas:
        beneficiario = vivienda.beneficiario
        ws.append([
            vivienda.proyecto.nombre if vivienda.proyecto else "",
            vivienda.codigo,
            vivienda.tipologia.nombre if hasattr(vivienda, 'tipologia') and vivienda.tipologia else "",
            vivienda.estado,
            beneficiario.nombre_completo if beneficiario else "",
            beneficiario.rut if beneficiario else "",
            beneficiario.email if beneficiario else ""
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_total.xlsx'
    wb.save(response)
    return response

from django.contrib.auth.decorators import login_required
from proyectos.models import Proyecto, Vivienda, Beneficiario

@login_required
def reporte_beneficiarios_por_proyecto_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Beneficiarios por Proyecto"
    headers = ["Proyecto", "Vivienda", "Beneficiario", "RUT", "Email"]
    ws.append(headers)
    proyectos = Proyecto.objects.filter(activo=True)
    for proyecto in proyectos:
        viviendas = Vivienda.objects.filter(proyecto=proyecto)
        for vivienda in viviendas:
            beneficiario = vivienda.beneficiario
            if beneficiario:
                ws.append([
                    proyecto.nombre,
                    vivienda.codigo,
                    beneficiario.nombre_completo,
                    beneficiario.rut,
                    beneficiario.email
                ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=beneficiarios_por_proyecto.xlsx'
    wb.save(response)
    return response

@login_required
def reporte_viviendas_sin_beneficiario_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Viviendas sin Beneficiario"
    headers = ["Proyecto", "Vivienda", "Tipología", "Estado"]
    ws.append(headers)
    viviendas = Vivienda.objects.filter(beneficiario__isnull=True, activa=True)
    for vivienda in viviendas:
        ws.append([
            vivienda.proyecto.nombre,
            vivienda.codigo,
            vivienda.tipologia.nombre,
            vivienda.estado
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=viviendas_sin_beneficiario.xlsx'
    wb.save(response)
    return response
@login_required
def reporte_observaciones_abiertas_urgentes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Observaciones Abiertas Urgentes"
    headers = ["ID", "Proyecto", "Vivienda", "Descripción", "Estado", "Urgente", "Fecha", "Usuario"]
    ws.append(headers)
    obs = Observacion.objects.filter(estado__nombre='Abierta', es_urgente=True)
    for o in obs:
        fecha = o.fecha_creacion.replace(tzinfo=None) if o.fecha_creacion else ""
        ws.append([
            o.id,
            getattr(o.vivienda.proyecto, 'nombre', ''),
            getattr(o.vivienda, 'codigo', ''),
            o.detalle,
            o.estado.nombre,
            o.es_urgente,
            fecha,
            getattr(o.creado_por, 'username', '')
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=observaciones_abiertas_urgentes.xlsx'
    wb.save(response)
    return response

@login_required
def reporte_observaciones_cerradas_urgentes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Observaciones Cerradas Urgentes"
    headers = ["ID", "Proyecto", "Vivienda", "Descripción", "Estado", "Urgente", "Fecha", "Usuario"]
    ws.append(headers)
    obs = Observacion.objects.filter(estado__nombre='Cerrada', es_urgente=True)
    for o in obs:
        fecha = o.fecha_creacion.replace(tzinfo=None) if o.fecha_creacion else ""
        ws.append([
            o.id,
            getattr(o.vivienda.proyecto, 'nombre', ''),
            getattr(o.vivienda, 'codigo', ''),
            o.detalle,
            o.estado.nombre,
            o.es_urgente,
            fecha,
            getattr(o.creado_por, 'username', '')
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=observaciones_cerradas_urgentes.xlsx'
    wb.save(response)
    return response

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from incidencias.models import Observacion

import openpyxl
from proyectos.models import Proyecto
@login_required
def reporte_estadisticas_region_excel(request):
    import openpyxl
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estadísticas por Región"
    headers = [
        "Región", "Proyecto", "Constructora", "Total Casas", "Casas con Observaciones", "Observaciones Abiertas", "Observaciones Urgentes", "Observaciones Cerradas", "% Abiertas", "% Urgentes", "% Cerradas"
    ]
    ws.append(headers)

    from proyectos.models import Proyecto, Vivienda
    from incidencias.models import Observacion
    regiones = Proyecto.objects.values_list('region', flat=True).distinct()
    for region_id in regiones:
        region_obj = None
        try:
            from core.models import Region
            region_obj = Region.objects.get(pk=region_id)
            region_nombre = region_obj.nombre
        except Exception:
            region_nombre = str(region_id)
        proyectos = Proyecto.objects.filter(region=region_id)
        for proyecto in proyectos:
            constructora = proyecto.constructora.nombre if proyecto.constructora else "-"
            total_casas = Vivienda.objects.filter(proyecto=proyecto).count()
            casas_con_obs = Vivienda.objects.filter(proyecto=proyecto, observaciones__isnull=False).distinct().count()
            obs = Observacion.objects.filter(vivienda__proyecto=proyecto)
            total_obs = obs.count()
            abiertas = obs.filter(estado__nombre='Abierta').count()
            urgentes = obs.filter(es_urgente=True, estado__nombre='Abierta').count()
            cerradas = obs.filter(estado__nombre='Cerrada').count()
            porc_abiertas = round((abiertas / total_obs) * 100, 1) if total_obs else 0
            porc_urgentes = round((urgentes / total_obs) * 100, 1) if total_obs else 0
            porc_cerradas = round((cerradas / total_obs) * 100, 1) if total_obs else 0
            ws.append([
                region_nombre,
                proyecto.nombre,
                constructora,
                total_casas,
                casas_con_obs,
                abiertas,
                urgentes,
                cerradas,
                porc_abiertas,
                porc_urgentes,
                porc_cerradas
            ])

    # Crear gráfico de barras para observaciones por región
    chart = BarChart()
    chart.title = "Observaciones por Región"
    chart.y_axis.title = "Cantidad"
    chart.x_axis.title = "Región"
    data = Reference(ws, min_col=3, max_col=5, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "N2")

    # Mejorar formato de encabezados
    for cell in ws[1]:
        cell.font = Font(bold=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=estadisticas_region_completo.xlsx'
    wb.save(response)
    return response

@login_required
def reporte_observaciones_cerradas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Observaciones Cerradas"
    headers = ["ID", "Proyecto", "Vivienda", "Descripción", "Estado", "Urgente", "Fecha", "Usuario"]
    ws.append(headers)
    obs = Observacion.objects.filter(estado__nombre='Cerrada', es_urgente=False)
    for o in obs:
        fecha = o.fecha_creacion.replace(tzinfo=None) if o.fecha_creacion else ""
        ws.append([
            o.id,
            getattr(o.vivienda.proyecto, 'nombre', ''),
            getattr(o.vivienda, 'codigo', ''),
            o.detalle,
            o.estado.nombre,
            o.es_urgente,
            fecha,
            getattr(o.creado_por, 'username', '')
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=observaciones_cerradas.xlsx'
    wb.save(response)
    return response

@login_required
def reporte_observaciones_abiertas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Observaciones Abiertas"
    headers = ["ID", "Proyecto", "Vivienda", "Descripción", "Estado", "Urgente", "Fecha", "Usuario"]
    ws.append(headers)
    obs = Observacion.objects.filter(estado__nombre='Abierta', es_urgente=False)
    for o in obs:
        fecha = o.fecha_creacion.replace(tzinfo=None) if o.fecha_creacion else ""
        ws.append([
            o.id,
            getattr(o.vivienda.proyecto, 'nombre', ''),
            getattr(o.vivienda, 'codigo', ''),
            o.detalle,
            o.estado.nombre,
            o.es_urgente,
            fecha,
            getattr(o.creado_por, 'username', '')
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=observaciones_abiertas.xlsx'
    wb.save(response)
    return response

@login_required
def reporte_observaciones_en_ejecucion_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Observaciones en Ejecución"
    headers = ["ID", "Proyecto", "Vivienda", "Detalle", "Estado", "Urgente", "Fecha Creación", "Usuario"]
    ws.append(headers)
    obs = Observacion.objects.filter(estado__nombre='En Ejecución')
    for o in obs:
        fecha = o.fecha_creacion.replace(tzinfo=None) if o.fecha_creacion else ""
        ws.append([
            o.id,
            getattr(o.vivienda.proyecto, 'nombre', ''),
            getattr(o.vivienda, 'codigo', ''),
            o.detalle,
            o.estado.nombre,
            o.es_urgente,
            fecha,
            getattr(o.creado_por, 'username', '')
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=observaciones_en_ejecucion.xlsx'
    wb.save(response)
    return response
    for o in obs:
        fecha = o.fecha_creacion.replace(tzinfo=None) if o.fecha_creacion else ""
        ws.append([
            o.id,
            getattr(o.vivienda.proyecto, 'nombre', ''),
            getattr(o.vivienda, 'codigo', ''),
            o.detalle,
            o.estado.nombre,
            o.es_urgente,
            fecha,
            getattr(o.creado_por, 'username', '')
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=observaciones_abiertas.xlsx'
    wb.save(response)
    return response

@login_required
def reporte_observaciones_urgentes_pendientes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Urgentes Pendientes"
    headers = ["ID", "Proyecto", "Vivienda", "Descripción", "Estado", "Urgente", "Fecha", "Usuario"]
    ws.append(headers)
    obs = Observacion.objects.filter(es_urgente=True, estado__nombre='Abierta')
    for o in obs:
        fecha = o.fecha_creacion.replace(tzinfo=None) if o.fecha_creacion else ""
        ws.append([
            o.id,
            getattr(o.vivienda.proyecto, 'nombre', ''),
            getattr(o.vivienda, 'codigo', ''),
            o.detalle,
            o.estado.nombre,
            o.es_urgente,
            fecha,
            getattr(o.creado_por, 'username', '')
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=urgentes_pendientes.xlsx'
    wb.save(response)
    return response

@login_required
def reporte_observaciones_urgentes_cerradas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Urgentes Cerradas"
    headers = ["ID", "Proyecto", "Vivienda", "Descripción", "Estado", "Urgente", "Fecha", "Usuario"]
    ws.append(headers)
    obs = Observacion.objects.filter(es_urgente=True, estado__nombre='Cerrada')
    for o in obs:
        fecha = o.fecha_creacion.replace(tzinfo=None) if o.fecha_creacion else ""
        ws.append([
            o.id,
            getattr(o.vivienda.proyecto, 'nombre', ''),
            getattr(o.vivienda, 'codigo', ''),
            o.detalle,
            o.estado.nombre,
            o.es_urgente,
            fecha,
            getattr(o.creado_por, 'username', '')
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=urgentes_cerradas.xlsx'
    wb.save(response)
    return response

@login_required
def reporte_observaciones_urgentes_abiertas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Urgentes Abiertas"
    headers = ["ID", "Proyecto", "Vivienda", "Descripción", "Estado", "Urgente", "Fecha", "Usuario"]
    ws.append(headers)
    obs = Observacion.objects.filter(es_urgente=True, estado__nombre='Abierta')
    for o in obs:
        fecha = o.fecha_creacion.replace(tzinfo=None) if o.fecha_creacion else ""
        ws.append([
            o.id,
            getattr(o.vivienda.proyecto, 'nombre', ''),
            getattr(o.vivienda, 'codigo', ''),
            o.detalle,
            o.estado.nombre,
            o.es_urgente,
            fecha,
            getattr(o.creado_por, 'username', '')
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=urgentes_abiertas.xlsx'
    wb.save(response)
    return response
import openpyxl
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
@login_required
def reporte_entregas_excel(request):
    """Exporta todas las actas de entrega y estadísticas en formato Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entregas"

    # Encabezados
    headers = [
        "N° Acta", "Fecha Entrega", "Proyecto", "Código Proyecto", "Comuna", "Región", "Beneficiario", "RUT", "Email", "Código Vivienda", "Tipología", "Estado Vivienda", "Familia Beneficiaria"
    ]
    ws.append(headers)

    # Datos
    # Filtros por proyecto y fechas
    proyecto_id = request.GET.get('proyecto')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    actas = ActaRecepcion.objects.select_related('proyecto', 'vivienda', 'beneficiario')
    if proyecto_id:
        actas = actas.filter(proyecto_id=proyecto_id)
    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            actas = actas.filter(fecha_entrega__gte=fecha_inicio_dt)
        except Exception:
            pass
    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
            actas = actas.filter(fecha_entrega__lte=fecha_fin_dt)
        except Exception:
            pass

    for acta in actas:
        ws.append([
            acta.numero_acta,
            acta.fecha_entrega.strftime('%d/%m/%Y') if acta.fecha_entrega else '',
            acta.proyecto.nombre if acta.proyecto else '',
            acta.proyecto.codigo if acta.proyecto else '',
            acta.proyecto.comuna.nombre if acta.proyecto and acta.proyecto.comuna else '',
            acta.proyecto.region.nombre if acta.proyecto and acta.proyecto.region else '',
            acta.beneficiario.nombre_completo if acta.beneficiario else '',
            acta.beneficiario.rut if acta.beneficiario else '',
            acta.beneficiario.email if acta.beneficiario else '',
            acta.vivienda.codigo if acta.vivienda else '',
            acta.vivienda.tipologia.nombre if acta.vivienda and acta.vivienda.tipologia else '',
            acta.vivienda.get_estado_display() if acta.vivienda else '',
            acta.vivienda.familia_beneficiaria if acta.vivienda else '',
        ])

    # Ajustar ancho de columnas
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 2)

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=entregas_techo_chile.xlsx'
    wb.save(response)
    return response
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.db import transaction
from datetime import datetime
import os
from django.conf import settings
from io import BytesIO

# El sistema ahora usa xhtml2pdf en lugar de WeasyPrint para compatibilidad con Windows
# WEASYPRINT_AVAILABLE = False  # deprecado - ahora se usa xhtml2pdf

# Fallback a ReportLab
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    # Definir valores por defecto si no está disponible
    inch = 72
    cm = 28.35

# Usaremos xhtml2pdf para generar PDFs (compatible con Windows); si falla, usamos ReportLab

from .models import ActaRecepcion, FamiliarBeneficiario  # , ConstructorActa
from proyectos.models import Vivienda, Proyecto, Beneficiario
from core.decorators import rol_requerido


@login_required
def index(request):
    """Vista principal de reportes"""
    context = {
        'titulo': 'Reportes y Documentos',
        'total_actas': ActaRecepcion.objects.count(),
        'actas_mes': ActaRecepcion.objects.filter(
            fecha_creacion__month=timezone.now().month,
            fecha_creacion__year=timezone.now().year
        ).count(),
    }
    return render(request, 'reportes/index.html', context)


@login_required
@rol_requerido('ADMINISTRADOR', 'TECHO')
def acta_list(request):
    """Lista de actas de recepción"""
    actas = ActaRecepcion.objects.select_related(
        'vivienda', 'proyecto', 'beneficiario'
    ).order_by('-fecha_creacion')
    
    # Filtros
    proyecto_id = request.GET.get('proyecto')
    if proyecto_id:
        actas = actas.filter(proyecto_id=proyecto_id)
    
    context = {
        'titulo': 'Actas de Recepción',
        'actas': actas,
        'proyectos': Proyecto.objects.filter(activo=True).order_by('nombre'),
    }
    return render(request, 'reportes/acta_list.html', context)


@login_required
@rol_requerido('ADMINISTRADOR', 'TECHO')
def acta_create(request):
    """Crear nueva acta de recepción"""
    from .forms import ActaRecepcionForm
    
    if request.method == 'POST':
        form = ActaRecepcionForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    acta = form.save(commit=False)
                    acta.save()
                    
                    messages.success(request, f'Acta {acta.numero_acta} creada exitosamente.')
                    return redirect('reportes:acta_detail', pk=acta.pk)
            except Exception as e:
                messages.error(request, f'Error al crear el acta: {str(e)}')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f'Error: {error}')
                    else:
                        messages.error(request, f'Error en {field}: {error}')
    else:
        form = ActaRecepcionForm()
    
    context = {
        'titulo': 'Crear Acta de Recepción',
        'form': form,
        'proyectos': Proyecto.objects.filter(activo=True).order_by('nombre'),
    }
    return render(request, 'reportes/acta_create.html', context)


@login_required
def acta_detail(request, pk):
    """Ver detalle de acta"""
    acta = get_object_or_404(
        ActaRecepcion.objects.select_related(
            'vivienda', 'proyecto', 'beneficiario'
        ).prefetch_related('familiares'),  # 'constructor_info'
        pk=pk
    )
    
    context = {
        'titulo': f'Acta {acta.numero_acta}',
        'acta': acta,
    }
    return render(request, 'reportes/acta_detail.html', context)


@login_required  
def acta_pdf(request, pk):
    """Generar PDF del acta - con opción de descarga"""
    acta = get_object_or_404(
        ActaRecepcion.objects.select_related(
            'vivienda', 'proyecto', 'beneficiario'  
        ).prefetch_related('familiares'),
        pk=pk
    )
    
    es_descarga = request.GET.get('download') == '1'
    
    # GENERAR PDF completo si download=1
    if es_descarga:
        # Ruta mínima de diagnóstico: generar un PDF básico si ?mode=min
        if request.GET.get('mode') == 'min':
            try:
                from reportlab.pdfgen import canvas  # type: ignore
                buffer = BytesIO()
                c = canvas.Canvas(buffer, pagesize=A4)
                c.drawString(100, 800, f"Acta {acta.numero_acta} - PDF minimal")
                c.drawString(100, 780, timezone.now().strftime('%d/%m/%Y %H:%M'))
                c.showPage()
                c.save()
                buffer.seek(0)
                response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="Acta_{acta.numero_acta}_minimal.pdf"'
                response['X-PDF-Engine'] = 'reportlab-min'
                return response
            except Exception as e:
                print(f"ERROR PDF minimal: {e}")
        # 1) Usar xhtml2pdf para compatibilidad con Windows
        try:
            from xhtml2pdf import pisa
            context = {
                'acta': acta,
                'proyecto': acta.proyecto,
                'beneficiario': acta.beneficiario,
                'vivienda': acta.vivienda,
                'familiares': acta.familiares.all(),
                'fecha_generacion': timezone.now(),
                'es_descarga': True,
                'para_pdf': True,
            }
            html_content = render_to_string('reportes/acta_template.html', context, request=request)
            pdf_buffer = BytesIO()
            pisa_status = pisa.CreatePDF(html_content, dest=pdf_buffer)
            if not pisa_status.err:
                pdf_file = pdf_buffer.getvalue()
                pdf_buffer.close()
                response = HttpResponse(pdf_file, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="Acta de Recepción - {acta.numero_acta}.pdf"'
                response['X-PDF-Engine'] = 'xhtml2pdf'
                return response
            pdf_buffer.close()
        except Exception:
            import traceback
            print("ERROR PDF xhtml2pdf:\n" + traceback.format_exc())
        # 2) Intentar con ReportLab (estable en Windows)
        # Usar ReportLab si está disponible a nivel de módulo

        if REPORTLAB_AVAILABLE:
            try:
                # Generar PDF usando ReportLab
                buffer = BytesIO()
                doc = SimpleDocTemplate(
                    buffer,
                    pagesize=A4,
                    rightMargin=2*cm,
                    leftMargin=2*cm,
                    topMargin=2*cm,
                    bottomMargin=2*cm,
                )

                # Estilos personalizados que coincidan con el HTML
                styles = getSampleStyleSheet()

                # Estilo para el encabezado principal
                header_style = ParagraphStyle(
                    'CustomHeader',
                    parent=styles['Title'],
                    fontSize=18,
                    spaceAfter=10,
                    alignment=1,  # Centrado
                    textColor=colors.black,
                    fontName='Helvetica-Bold'
                )

                # Estilo para subtítulo
                subtitle_style = ParagraphStyle(
                    'CustomSubtitle',
                    parent=styles['Heading2'],
                    fontSize=14,
                    spaceAfter=20,
                    alignment=1,
                    textColor=colors.black,
                    fontName='Helvetica-Bold'
                )

                # Estilo para secciones
                section_style = ParagraphStyle(
                    'SectionHeader',
                    parent=styles['Heading3'],
                    fontSize=12,
                    spaceBefore=15,
                    spaceAfter=10,
                    textColor=colors.black,
                    fontName='Helvetica-Bold'
                )

                # Estilo para texto normal
                normal_style = ParagraphStyle(
                    'CustomNormal',
                    parent=styles['Normal'],
                    fontSize=10,
                    spaceAfter=4,
                    fontName='Helvetica'
                )

                story = []

                # === ENCABEZADO ===
                story.append(Paragraph("ACTA DE RECEPCIÓN DE VIVIENDA", header_style))
                story.append(Paragraph("TECHO CHILE", subtitle_style))
                story.append(Paragraph(f"<b>Acta N°:</b> {acta.numero_acta} | <b>Generada:</b> {timezone.now().strftime('%d/%m/%Y %H:%M')}", normal_style))
                story.append(Spacer(1, 20))

                # === INFORMACIÓN DEL PROYECTO ===
                story.append(Paragraph("INFORMACIÓN DEL PROYECTO", section_style))
                proyecto_data = [
                    ['Proyecto:', acta.proyecto.nombre if acta.proyecto else 'N/A'],
                    ['Comuna:', acta.proyecto.comuna.nombre if (acta.proyecto and acta.proyecto.comuna) else 'N/A'],
                    ['Región:', acta.proyecto.region.nombre if (acta.proyecto and acta.proyecto.region) else 'N/A'],
                ]

                for label, value in proyecto_data:
                    story.append(Paragraph(f"<b>{label}</b> {value}", normal_style))
                story.append(Spacer(1, 12))

                # === INFORMACIÓN DEL BENEFICIARIO ===
                story.append(Paragraph("INFORMACIÓN DEL BENEFICIARIO", section_style))
                if acta.beneficiario:
                    # Construir string de teléfonos activos
                    telefonos_activos = list(acta.beneficiario.telefonos.filter(activo=True).values_list('numero', flat=True))
                    if not telefonos_activos:
                        telefonos_activos = list(acta.beneficiario.telefonos.values_list('numero', flat=True)[:1])
                    telefonos_str = ', '.join(telefonos_activos) if telefonos_activos else 'N/A'

                    beneficiario_data = [
                        ['Nombre Completo:', f"{acta.beneficiario.nombre} {acta.beneficiario.apellido_paterno} {acta.beneficiario.apellido_materno or ''}".strip()],
                        ['RUT:', acta.beneficiario.rut or 'N/A'],
                        ['Teléfonos:', telefonos_str],
                        ['Email:', acta.beneficiario.email or 'N/A'],
                    ]

                    for label, value in beneficiario_data:
                        story.append(Paragraph(f"<b>{label}</b> {value}", normal_style))
                else:
                    story.append(Paragraph("<b>Beneficiario:</b> No asignado", normal_style))
                story.append(Spacer(1, 12))

                # === INFORMACIÓN DE LA VIVIENDA ===
                story.append(Paragraph("INFORMACIÓN DE LA VIVIENDA", section_style))
                if acta.vivienda:
                    vivienda_data = [
                        ['Código de Vivienda:', acta.vivienda.codigo],
                        ['Tipología:', acta.vivienda.tipologia.nombre if acta.vivienda.tipologia else 'N/A'],
                        ['Estado:', acta.vivienda.get_estado_display()],
                        ['Familia Beneficiaria:', acta.vivienda.familia_beneficiaria or 'N/A'],
                    ]

                    for label, value in vivienda_data:
                        story.append(Paragraph(f"<b>{label}</b> {value}", normal_style))
                else:
                    story.append(Paragraph("<b>Vivienda:</b> No asignada", normal_style))
                story.append(Spacer(1, 12))

                # === DETALLES DEL ACTA ===
                story.append(Paragraph("DETALLES DEL ACTA", section_style))
                acta_data = [
                    ['Fecha de Entrega:', acta.fecha_entrega.strftime("%d/%m/%Y") if acta.fecha_entrega else 'N/A'],
                    ['Entrega Conforme:', 'Sí' if acta.entregado_beneficiario else 'No'],
                    ['Estructura:', 'Conforme' if getattr(acta, 'estado_estructura', True) else 'Con observaciones'],
                    ['Instalaciones:', 'Conforme' if getattr(acta, 'estado_instalaciones', True) else 'Con observaciones'],
                ]

                for label, value in acta_data:
                    story.append(Paragraph(f"<b>{label}</b> {value}", normal_style))
                story.append(Spacer(1, 15))

                # === FAMILIARES (si los hay) ===
                familiares = acta.familiares.all()
                if familiares.exists():
                    story.append(Paragraph("FAMILIARES REGISTRADOS", section_style))

                    # Crear tabla de familiares
                    familiares_data = [['Nombre Completo', 'Parentesco', 'RUT', 'Edad']]
                    for familiar in familiares:
                        familiares_data.append([
                            familiar.nombre_completo,
                            familiar.parentesco,
                            familiar.rut or 'N/A',
                            str(familiar.edad) if familiar.edad else 'N/A'
                        ])

                    familiares_table = Table(familiares_data, colWidths=[8*cm, 3*cm, 3*cm, 2*cm])
                    familiares_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ]))

                    story.append(familiares_table)
                    story.append(Spacer(1, 15))

                # === OBSERVACIONES ===
                if acta.observaciones:
                    story.append(Paragraph("OBSERVACIONES", section_style))
                    story.append(Paragraph(acta.observaciones, normal_style))
                    story.append(Spacer(1, 20))

                # === FIRMAS ===
                story.append(Spacer(1, 30))
                story.append(Paragraph("FIRMAS Y CONFORMIDAD", section_style))

                # Tabla de firmas
                firma_data = [
                    ['', '', ''],
                    ['_________________________', '_________________________', '_________________________'],
                    ['Beneficiario', 'Representante TECHO', 'Supervisor'],
                    ['', '', ''],
                    ['Fecha: _______________', 'Fecha: _______________', 'Fecha: _______________']
                ]

                firma_table = Table(firma_data, colWidths=[5.5*cm, 5.5*cm, 5.5*cm])
                firma_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
                    ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ]))

                story.append(firma_table)

                # Construir PDF
                doc.build(story)

                # Preparar respuesta PDF
                buffer.seek(0)
                response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="Acta de Recepción - {acta.numero_acta}.pdf"'
                response['X-PDF-Engine'] = 'reportlab'

                return response

            except Exception as e:
                # Log del error pero continuar con HTML
                import traceback
                print("ERROR PDF ReportLab:\n" + traceback.format_exc())

        # (xhtml2pdf y ReportLab ya fueron intentados arriba)

        # 3) Si ambos fallaron, generar un PDF mínimo como último recurso
        try:
            from reportlab.pdfgen import canvas  # type: ignore
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=A4)
            c.setFont('Helvetica', 12)
            c.drawString(72, 800, f"Acta {acta.numero_acta}")
            c.drawString(72, 780, "No se pudo generar el PDF completo. Este es un PDF mínimo de respaldo.")
            c.drawString(72, 760, timezone.now().strftime('%d/%m/%Y %H:%M'))
            c.showPage()
            c.save()
            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="Acta de Recepción - {acta.numero_acta} (respaldo).pdf"'
            response['X-PDF-Engine'] = 'reportlab-min-fallback'
            return response
        except Exception:
            return HttpResponse("No se pudo generar el PDF en este momento.", status=500, content_type='text/plain; charset=utf-8')
    
    # HTML por defecto (vista previa)
    context = {
        'acta': acta,
        'proyecto': acta.proyecto,
        'beneficiario': acta.beneficiario,
        'vivienda': acta.vivienda,
        'familiares': acta.familiares.all(),
        'fecha_generacion': timezone.now(),
        'es_descarga': es_descarga,
    }
    
    html_content = render_to_string('reportes/acta_template.html', context)
    
    if es_descarga:
        # Si no logramos generar PDF, mostramos HTML sin forzar descarga binaria
        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = f'inline; filename="Acta_{acta.numero_acta}_{timezone.now().strftime("%Y%m%d")}.html"'
        response['X-PDF-Engine'] = 'html-fallback'
    else:
        response = HttpResponse(html_content, content_type='text/html')
        response['X-PDF-Engine'] = 'html-view'
    
    return response


@login_required
def get_viviendas_by_proyecto(request):
    """AJAX: Obtener viviendas por proyecto"""
    proyecto_id = request.GET.get('proyecto_id')
    viviendas = Vivienda.objects.filter(
        proyecto_id=proyecto_id,
        activa=True,
        beneficiario__isnull=False
    ).select_related('beneficiario').values(
        'id', 'codigo', 'beneficiario__nombre_completo'
    )
    
    return JsonResponse(list(viviendas), safe=False)


@login_required
def get_vivienda_data(request):
    """AJAX: Obtener datos de la vivienda para prellenar el formulario"""
    vivienda_id = request.GET.get('vivienda_id')
    vivienda = get_object_or_404(
        Vivienda.objects.select_related(
            'beneficiario', 'proyecto', 'proyecto__constructora', 'tipologia'
        ), 
        id=vivienda_id
    )
    
    # Obtener teléfono del beneficiario
    telefono = ""
    if vivienda.beneficiario and vivienda.beneficiario.telefonos.exists():
        telefono = vivienda.beneficiario.telefonos.filter(activo=True).first().numero
    
    # Mapear tipo estructura desde tipología si existe
    tipos_validos = {'madera', 'metalica', 'mixta', 'hormigon'}
    tipo_estructura_tipologia = getattr(vivienda.tipologia, 'tipo_estructura', None) if vivienda.tipologia else None
    tipo_estructura_res = tipo_estructura_tipologia if (isinstance(tipo_estructura_tipologia, str) and tipo_estructura_tipologia.lower() in tipos_validos) else None

    data = {
        'beneficiario': {
            'nombre_completo': vivienda.beneficiario.nombre_completo if vivienda.beneficiario else '',
            'rut': vivienda.beneficiario.rut if vivienda.beneficiario else '',
            'email': vivienda.beneficiario.email if vivienda.beneficiario else '',
            'telefono': telefono,
        },
        'proyecto': {
            'nombre': vivienda.proyecto.nombre,
            'comuna': vivienda.proyecto.comuna.nombre,
            'region': vivienda.proyecto.region.nombre,
        },
        'constructora': {
            'nombre': vivienda.proyecto.constructora.nombre,
            'rut': vivienda.proyecto.constructora.rut or '',
            'telefono': vivienda.proyecto.constructora.telefono or '',
        } if vivienda.proyecto.constructora else None,
        'vivienda': {
            'codigo': vivienda.codigo,
            'direccion': f"{vivienda.proyecto.nombre} - Vivienda {vivienda.codigo}",
            'superficie': getattr(vivienda.tipologia, 'metros_cuadrados', None),
            'ambientes': getattr(vivienda.tipologia, 'numero_ambientes', None),
            'tipo_estructura': tipo_estructura_res,
        }
    }
    
    return JsonResponse(data)


@login_required
@rol_requerido('ADMINISTRADOR', 'TECHO')
def acta_edit(request, pk):
    """Editar acta de recepción"""
    from .forms import ActaRecepcionForm
    
    acta = get_object_or_404(ActaRecepcion, pk=pk)
    
    if request.method == 'POST':
        form = ActaRecepcionForm(request.POST, instance=acta)
        if form.is_valid():
            try:
                acta = form.save()
                messages.success(request, f'Acta {acta.numero_acta} actualizada exitosamente.')
                return redirect('reportes:acta_detail', pk=acta.pk)
            except Exception as e:
                messages.error(request, f'Error al actualizar el acta: {str(e)}')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f'Error: {error}')
                    else:
                        messages.error(request, f'Error en {field}: {error}')
    else:
        form = ActaRecepcionForm(instance=acta)
    
    context = {
        'titulo': f'Editar Acta {acta.numero_acta}',
        'form': form,
        'acta': acta,
        'proyectos': Proyecto.objects.filter(activo=True).order_by('nombre'),
    }
    return render(request, 'reportes/acta_edit.html', context)


@login_required
@rol_requerido('ADMINISTRADOR', 'TECHO')
def acta_delete(request, pk):
    """Eliminar acta de recepción"""
    acta = get_object_or_404(ActaRecepcion, pk=pk)
    
    if request.method == 'POST':
        try:
            numero_acta = acta.numero_acta
            acta.delete()
            messages.success(request, f'Acta {numero_acta} eliminada exitosamente.')
            return redirect('reportes:acta_list')
        except Exception as e:
            messages.error(request, f'Error al eliminar el acta: {str(e)}')
            return redirect('reportes:acta_detail', pk=pk)
    
    # Si no es POST, redirigir al detalle
    return redirect('reportes:acta_detail', pk=pk)


@login_required
def buscar_beneficiario_ajax(request):
    """Vista AJAX para buscar datos del beneficiario por RUT"""
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Solicitud no válida'})
    
    rut = request.GET.get('rut', '').strip()
    if not rut:
        return JsonResponse({'success': False, 'error': 'RUT requerido'})
    
    try:
        # Buscar beneficiario por RUT
        beneficiario = Beneficiario.objects.select_related().get(rut=rut, activo=True)
        
        # Obtener datos de la vivienda asociada
        vivienda = None
        proyecto = None
        observaciones_previas = ""
        
        try:
            vivienda = beneficiario.vivienda
            if vivienda:
                proyecto = vivienda.proyecto
                observaciones_previas = vivienda.observaciones_generales or ""
        except:
            pass
        
        # Obtener observaciones de incidencias si existen
        try:
            from incidencias.models import Observacion
            observaciones_incidencias = Observacion.objects.filter(
                vivienda=vivienda,
                activo=True
            ).select_related('estado').values_list('descripcion', 'estado__nombre', 'es_urgente')
            
            if observaciones_incidencias:
                obs_lines = []
                obs_lines.append("OBSERVACIONES DE LA VIVIENDA:")
                obs_lines.append("=" * 40)
                
                for desc, estado, es_urgente in observaciones_incidencias:
                    plazo = "48 HORAS" if es_urgente else "120 DÍAS"
                    obs_lines.append(f"• {desc}")
                    obs_lines.append(f"  Estado: {estado} | Plazo: {plazo}")
                    obs_lines.append("")
                
                obs_texto = "\n".join(obs_lines)
                
                if observaciones_previas:
                    observaciones_previas = f"{observaciones_previas}\n\n{obs_texto}"
                else:
                    observaciones_previas = obs_texto
        except Exception as e:
            # Si hay error con incidencias, al menos incluir observaciones generales
            print(f"Error obteniendo observaciones: {e}")
            if not observaciones_previas and vivienda:
                observaciones_previas = vivienda.observaciones_generales or ""
        
        # Preparar datos de respuesta
        # Mapear tipo de estructura de tipología al choice del Acta (fallback 'madera')
        tipo_estructura_tipologia = None
        if vivienda and vivienda.tipologia:
            tipo_estructura_tipologia = getattr(vivienda.tipologia, 'tipo_estructura', None)
        tipos_validos = {'madera', 'metalica', 'mixta', 'hormigon'}
        tipo_estructura_res = tipo_estructura_tipologia if (isinstance(tipo_estructura_tipologia, str) and tipo_estructura_tipologia.lower() in tipos_validos) else 'madera'

        data = {
            'success': True,
            'beneficiario': {
                'id': beneficiario.id,
                'nombre_completo': beneficiario.nombre_completo,
                'rut': beneficiario.rut,
                'email': beneficiario.email or '',
                'proyecto_id': proyecto.id if proyecto else None,
                'proyecto_nombre': str(proyecto) if proyecto else '',
                'vivienda_id': vivienda.id if vivienda else None,
                'vivienda_codigo': vivienda.codigo if vivienda else '',
                'observaciones': observaciones_previas,
                'direccion_proyecto': f"{proyecto.nombre}, {proyecto.comuna.nombre}" if proyecto else '',
                'vivienda': {
                    # Nombres de claves esperadas por el front
                    'superficie': getattr(vivienda.tipologia, 'metros_cuadrados', None) or 18,
                    'tipo_estructura': tipo_estructura_res,
                    'ambientes': getattr(vivienda.tipologia, 'numero_ambientes', None) or 2,
                    'codigo': vivienda.codigo if vivienda else '',
                    'tipologia': vivienda.tipologia.nombre if vivienda and vivienda.tipologia else 'Vivienda Básica',
                } if vivienda else {
                    'superficie': 18,
                    'ambientes': 2,
                    'tipo_estructura': 'madera'
                }
            }
        }
        
        return JsonResponse(data)
        
    except Beneficiario.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': f'No se encontró beneficiario con RUT {rut}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error al buscar beneficiario: {str(e)}'
        })